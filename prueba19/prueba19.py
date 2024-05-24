# Librerías para la Manipulación de Archivos y Datos
import sys  
from io import BytesIO  
import xlwings as xw  
import pandas as pd
from openpyxl import load_workbook
import os
import shutil
from PyPDF2 import PdfMerger 

# Librerías para la Creación y Manipulación de Gráficos
import matplotlib.pyplot as plt  
import networkx as nx  

# Librerías para la Interfaz Gráfica
import customtkinter as ctk  
from tkinter import filedialog, messagebox 
from tkinter import Tk, filedialog

""" # Ocultar la CMD de Windows al ejecutar el .exe
hide = win32gui.GetForegroundWindow()
win32gui.ShowWindow(hide, win32con.SW_HIDE) """

# Clase para redirigir stdout a un widget Text de Tkinter
class RedirectStdout:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(ctk.END, string)
        self.text_widget.see(ctk.END)

    def flush(self):
        pass

""" # Función que muestra una ventana emergente con un ícono de advertencia
def show_warning(message):
    ctk.messagebox.showwarning("Advertencia", message) """

# Función para seleccionar un archivo Excel
def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
    if not file_path:
        print("No se seleccionó ningún archivo. El programa se cerrará.")
        return
    else:
        file_entry.delete(0, ctk.END)
        file_entry.insert(0, file_path)
        excel_intermedio(file_path)

# Función para duplicar un archivo Excel
def excel_intermedio(original_file):
    try:
        # Determinar la ruta para la copia del archivo
        directory = os.path.dirname(original_file)
        filename = os.path.basename(original_file)
        new_file_path = os.path.join(directory, "copia_" + filename)

        # Verificar si ya existe una copia del archivo
        if os.path.exists(new_file_path):
            overwrite_confirmation = messagebox.askyesno("Advertencia", "Ya existe una copia del archivo. ¿Desea sobrescribirlo?")
            if not overwrite_confirmation:
                print("Operación cancelada por el usuario.")
                return

        # Copiar el archivo Excel original a la nueva ubicación
        shutil.copy(original_file, new_file_path)
        print(f'Archivo duplicado exitosamente en: {new_file_path}')

        # Llamar a la función de filtros_excel
        filtros_excel(new_file_path)

    except Exception as e:
        print(f'Ocurrió un error al guardar el Excel Intermedio')

# Función para filtrar cambios del archivo Excel
def filtros_excel(file_path):
    try:

       # Cargar el archivo Excel copiado
        workbook = load_workbook(filename=file_path)

        # Obtener la hoja activa
        sheet = workbook.active

        # Obtener el rango de celdas de la hoja
        max_row = sheet.max_row
        max_column = sheet.max_column

        ## Eliminar duplicados de celdas
        for column in range(1, max_column + 1):# Iterar sobre cada celda de cada columna (excepto la columna 6)
            if column == 6:  # Ignorar la columna 6
                continue
            for row in range(2, max_row + 1):
                cell = sheet.cell(row=row, column=column)
                value = cell.value
                if value and isinstance(value, str):

                    # Dividir el contenido de la celda en partes separadas por espacios
                    parts = value.split()

                    # Conservar solo el primer valor único
                    unique_values = set()
                    unique_parts = []
                    for part in parts:
                        if part not in unique_values:
                            unique_values.add(part)
                            unique_parts.append(part)

                    # Volver a unir las partes en un solo texto
                    new_value = ' '.join(unique_parts)

                    # Asignar el nuevo valor a la celda
                    cell.value = new_value
                
        print(f'Se han eliminado los duplicados correctamente en: {file_path}')

        ## Reemplazar "/" por "/ + \n" en la columna 6
        for row in range(2, max_row + 1):
            cell = sheet.cell(row=row, column=6)
            value = cell.value
            if value and isinstance(value, str):
                # Reemplazar "/ " por "/\n"
                new_value = value.replace("/", "/\n")
                # Asignar el nuevo valor a la celda
                cell.value = new_value

        print(f'Se han realizado los saltos de línea correctamente en: {file_path}')

        # Ignorar filas con 4 o más celdas vacías
        rows_to_delete = [] # Inicializar una lista para almacenar los índices de las filas a eliminar

        # Iterar sobre cada fila
        for row in range(2, max_row + 1):
            empty_cell_count = 0
            for column in range(1, max_column + 1):
                cell = sheet.cell(row=row, column=column)
                if cell.value is None:
                    empty_cell_count += 1
                else:
                    if empty_cell_count == 4:
                        rows_to_delete.pop()  # Eliminar el último índice si no son celdas vacías seguidas
                    empty_cell_count = 0  # Reiniciar el contador si se encuentra una celda no vacía
                if empty_cell_count >= 4:
                    rows_to_delete.append(row)  # Agregar el índice de la fila a la lista
                    break  # Salir del bucle si hay 4 o más celdas vacías seguidas
        
        # Eliminar las filas que contienen 4 o más celdas vacías seguidas
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)

        print(f'Se han eliminado las filas que no contienen información')

        # Guardar los cambios en el archivo Excel copiado
        workbook.save(filename=file_path)

        # Mostrar mensaje de confirmación si se sobrescribe el archivo
        messagebox.showinfo("Información", "Se han guardado los cambios en la copia del archivo.")

        # Abrir el archivo Excel copiado para visualización
        os.system(file_path) 

    except Exception as e:
        print(f'Ocurrió un error al filtrar datos del Excel.')
        raise e

# Función para procesar el archivo seleccionado
def process_file():
    original_file = file_entry.get()
    if original_file:
        excel_intermedio(original_file)
    else:
        print("No se seleccionó ningún archivo. El programa se cerrará.")


""" 
# Función para verificar la extensión del archivo Excel
def is_valid_extension(file_path, valid_extensions=('.xls', '.xlsx')):
    return file_path.endswith(valid_extensions)

# Función para guardar el archivo PDF combinado
def save_combined_pdf(pdf_merger):
    pdf_combined_file = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
    
    if not pdf_combined_file:
        show_warning("No se seleccionó ninguna ruta para guardar el archivo. El programa se cerrará.") # El programa se cierra si no se selecciona una ruta
        root.destroy()
    else:
        with open(pdf_combined_file, 'wb') as output_pdf:
            pdf_merger.write(output_pdf)
        pdf_merger.close()  # Cerrar el PdfMerger para liberar recursos

        # Mensaje de Información al finalizar el proceso
        print(f"El archivo PDF se ha guardado correctamente en: {pdf_combined_file}")
        ctk.messagebox.showinfo("Información", f"El archivo PDF se ha generado correctamente en: {pdf_combined_file}\n")

# Función para procesar el archivo seleccionado
def process_file():
    file_path = file_entry.get()
    tamaño_hoja = size_var.get()
    figsize, x_position_max = get_figsize_and_max_pos(tamaño_hoja)

    # Crear un objeto para combinar PDFs
    pdf_merger = PdfMerger()

    # Abrir el archivo de Excel con xlwings
    app = xw.App(visible=False)
    wb = app.books.open(file_path)  # Abrir el archivo seleccionado por el usuario
    sheet = wb.sheets[0]

    # Leer datos desde Excel
    df = sheet.used_range.value

    # Iterar sobre las filas del DataFrame
    for idx, row in enumerate(df, start=1):
        
        # Verificar si la fila tiene suficientes valores para dibujar el gráfico
        if len(row) > 1:  # Por lo menos dos nodos para crear una conexión

            # Crear un nuevo grafo para cada fila
            G = create_graph_from_row()

            # Dibujar el gráfico con el tamaño adecuado
            fig, ax = plt.subplots(figsize=figsize)

            # Calcular posiciones de los nodos
            pos = generate_positions(G, x_position_max)

            draw_graph(G, pos, ax)

            # Guardar la figura en un objeto BytesIO
            buf = BytesIO()
            plt.savefig(buf, format='pdf')
            buf.seek(0)
            pdf_merger.append(buf)
            buf.close()  # Cerrar BytesIO para liberar memoria
            plt.close(fig)  # Cerrar la figura para liberar memoria

            print(f"El diagrama de unifilares para la fila {idx} se ha generado con éxito\n")
        else:
            print(f"No ha sido posible hacer un diagrama para la fila {idx} porque no hay suficientes datos.\n")

    # Guardar el archivo PDF combinado
    save_combined_pdf(pdf_merger)

# Función para obtener el tamaño de la figura y la posición máxima en X
def get_figsize_and_max_pos(tamaño_hoja):
    if tamaño_hoja == 'A4': 
        return (210 / 25.4, 297 / 25.4), 4 # Tamaño A4
    else:                   
        return (420 / 25.4, 297 / 25.4), 8 # Tamaño A3

# Función para crear un grafo a partir de una fila
def create_graph_from_row(row):

    # Crear un nuevo grafo G
    G = nx.Graph() 
    
    # Agregar nodos al grafo G
    for cell in row:
        G.add_node(cell)

    # Agregar conexiones entre nodos en la fila actual
    for i in range(len(row) - 1):
        G.add_edge(row[i], row[i + 1])
    return G

# Función para generar posiciones de los nodos
def generate_positions(G, x_position_max):
    # Inicializar el diccionario de posiciones
    pos = {} 
    x_position, y_position = 0, 0

    for node in G.nodes():
        pos[node] = (x_position, y_position) # Inicializar el diccionario de posiciones

        # Posición de los nodos
        if x_position < x_position_max:
            x_position += 1
        else:
            x_position = 1
            y_position -= 1

    return pos

# Función para dibujar el grafo
def draw_graph(G, pos, ax):
    # Dibujar los nodos
    node_colors = ['skyblue' if i % 2 == 0 else 'lightgreen' for i in range(len(G.nodes))]
    node_shapes = ['s' if i % 2 == 0 else 'd' for i in range(len(G.nodes))]

    for i, (node, (x, y)) in enumerate(pos.items()): # Aquí, "pos.items()" devuelve una vista de los elementos (clave, valor) de "pos".
        nx.draw_networkx_nodes(G, pos, nodelist=[node], node_size=3000, node_shape=node_shapes[i], node_color=node_colors[i])
        ax.text(x, y, node, ha='center', va='center')

    nx.draw_networkx_edges(G, pos)
    ax.axis('off') """

###### Configuración de la apariencia de la ventana principal con customtkinter #####

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

root = ctk.CTk() # Crear la ventana principal
root.title("Generador de Diagramas Unifilares")
root.resizable(False, False)  # Desactivar redimensionamiento

# Botón 'Examinar'
ctk.CTkLabel(root, text="Seleccionar archivo de Excel:").grid(row=0, column=0, padx=10, pady=10)
file_entry = ctk.CTkEntry(root, width=300)
file_entry.grid(row=0, column=1, padx=10, pady=10)
ctk.CTkButton(root, text="Examinar", command=select_file).grid(row=0, column=2, padx=10, pady=10)

""" # Botón 'Tamaño hoja'
ctk.CTkLabel(root, text="Seleccionar tamaño de hoja:").grid(row=1, column=0, padx=10, pady=10)
size_var = ctk.StringVar(value="A4")
size_combobox = ctk.CTkComboBox(root, variable=size_var, values=["A4", "A3"])
size_combobox.grid(row=1, column=1, padx=10, pady=10) """


text_widget = ctk.CTkTextbox(root, wrap='word', height=200, width=600)
text_widget.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

sys.stdout = RedirectStdout(text_widget)

""" # Botón 'Generar PDF'
ctk.CTkButton(root, text="Generar PDF", command=process_file).grid(row=3, column=0, columnspan=3, pady=10)
 """
# Botón 'Salir'
ctk.CTkButton(root, text="Salir", command=root.quit).grid(row=4, column=0, columnspan=3, pady=10)

# Mantener la ventana abierta y a la espera de eventos (clics de ratón, pulsaciones de teclas, etc.)
root.mainloop()
