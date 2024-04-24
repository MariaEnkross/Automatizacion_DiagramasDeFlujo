import openpyxl
from datetime import datetime
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import landscape, A4

# Abre el archivo Excel lee solo los valores resultantes (ignorando las fórmulas)
wb = openpyxl.load_workbook('excel.xlsx', data_only=True)
sheet = wb['Tiempos']  # Se puede acceder a una hoja específica si es necesario

# Imprimir los nombres de todas las hojas en el archivo Excel
print("La hojas de este Excel son: ", wb.sheetnames)

# Lee datos del Excel y los almacena en una lista
data = []
for row in sheet.iter_rows(values_only=True):
    filtered_row = [value.strftime('%d-%m-%Y') 
                    if isinstance(value, datetime) 
                        else round(value, 1) if isinstance(value, (int, float)) 
                        else value for value in row if value is not None]  # Formatear fechas, redondear horas y filtrar celdas en blanco
    if filtered_row:  # Comprueba si filtered_row contiene algún valor despues del filtrado
        data.append(filtered_row)

# Crea un PDF
pdf_filename = "prueba3_datos_excel.pdf" 
doc = SimpleDocTemplate(pdf_filename, pagesize=landscape(A4))

# Crear una tabla con los datos
table = Table(data)

# Aplicar estilo a la tabla
style = TableStyle([('BACKGROUND', (0,0), (-1,0), colors.grey),
                    ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND',(0,1),(-1,-1),colors.beige),
                    ('GRID',(0,0),(-1,-1),1,colors.black)])
table.setStyle(style)

# Agregar la tabla al documento PDF
elements = [table]
doc.build(elements)

print(f"El PDF ha sido generado correctamente")

