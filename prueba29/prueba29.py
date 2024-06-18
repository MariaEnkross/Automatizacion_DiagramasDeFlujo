# Librerías para la Manipulación de Archivos y Datos
import sys
from io import BytesIO
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os
import shutil
from PyPDF2 import PdfMerger 

# Librerías para la Creación y Manipulación de Gráficos
import matplotlib.pyplot as plt  
import networkx as nx  
from reportlab.lib.pagesizes import A4, A3
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

# Librerías para la Interfaz Gráfica
import customtkinter as ctk  
from tkinter import filedialog, messagebox 


""" # Ocultar la CMD de Windows al ejecutar el .exe
hide = win32gui.GetForegroundWindow()
win32gui.ShowWindow(hide, win32con.SW_HIDE) """

# Clase para redirigir stdout a un widget Text de Tkinter
class RedirectStdout:

    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(ctk.END, string)
        self.text_widget.see(ctk.END)

    def flush(self):
        pass

# Variable global para almacenar la ruta del archivo procesado
new_file_path = ""

# Función para seleccionar un archivo Excel
def select_file():
    global new_file_path  # Asegurar que estamos utilizando la variable global

    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])

    if file_path:
        file_entry.delete(0, ctk.END)
        file_entry.insert(0, file_path)

        # Llamar a excel_intermedio con la ruta seleccionada
        new_file_path = excel_intermedio(file_path)
        if new_file_path:
            print(f"Ruta del archivo copiado: {new_file_path}")
        else:
            print("No se seleccionó ningún archivo. El programa se cerrará.")

def excel_intermedio(original_file):
    
    global new_file_path  # Asegurar que estamos utilizando la variable global

    try:
        # Determinar la ruta para la copia del archivo
        directory = os.path.dirname(original_file)
        filename = os.path.basename(original_file)
        new_file_path = os.path.join(directory, "copia_" + filename)

        # Verificar si ya existe una copia del archivo
        if os.path.exists(new_file_path):
            overwrite_confirmation = messagebox.askyesno("Advertencia", "Ya existe una copia del archivo. ¿Desea sobrescribirlo?")
            if not overwrite_confirmation:
                print("Operación cancelada por el usuario.")
                return

        # Copiar el archivo Excel original a la nueva ubicación
        shutil.copy(original_file, new_file_path)

        # Abrir el archivo Excel copiado con xlwings
        app = xw.App(visible=False)
        wb = app.books.open(new_file_path)

        # Cargar el archivo Excel copiado
        workbook = load_workbook(filename=new_file_path)

        # Función para configurar todas las celdas de las hojas al formato de texto.   
        def formato_texto(sheet):
            for row in sheet.iter_rows():
                for cell in row:
                    cell.number_format = '@'

        ## 1. Función para crear la hoja de errores en el archivo Excel copiado
        def hoja_errores(workbook):
            # Crear hoja llamada 'errores'
            error_sheet = workbook.create_sheet('errores')
            # Agregar las cabeceras de las columnas
            error_sheet['A1'] = 'Nombre de la Manguera:'
            error_sheet['B1'] = 'Página dónde se encuentra:'
            error_sheet['C1'] = 'Motivo del error:'
            formato_texto(error_sheet) # Asegurar que todas las celdas de la hoja de errores estén en formato texto

        # Llamar a la función para crear la hoja de errores
        hoja_errores(workbook)
            
        ## 2. Función para crear la hoja de uniones en el archivo Excel copiado
        def hoja_uniones(workbook):
            # Crear hoja llamada 'uniones'
            union_sheet = workbook.create_sheet('uniones')
            formato_texto(union_sheet) # Asegurar que todas las celdas de la hoja de uniones estén en formato texto

        # Llamar a la función para crear la hoja de uniones
        hoja_uniones(workbook)

        ## 3. Función para crear la hoja_mangueras_individuales en el archivo Excel copiado
        def hoja_mangueras_individuales(workbook):
            # Crear hoja llamada 'mangueras_individuales'
            mangueras_sheet = workbook.create_sheet('mangueras_individuales')
            formato_texto(mangueras_sheet) # Asegurar que todas las celdas de la hoja de mangueras estén en formato texto

        # Llamar a la función para crear la hoja de mangueras_individuales
        hoja_mangueras_individuales(workbook)

        # Guardar los cambios preliminares en el archivo copiado
        workbook.save(new_file_path)

        # Llamar a las funciones de filtros
        filtros_excel(new_file_path) # Llamar a la función de filtros_excel
        filtros_uniones(new_file_path) # Llamar a la función de filtros_uniones
        filtros_mangueras_individuales(new_file_path) # Llamar a la función de filtros_mangueras_individuales

        # Guardar los cambios
        wb.save()
        wb.close()
        app.quit()

        return new_file_path  # Devolver new_file_path después de procesar

    except Exception as e:
        print(f'Ocurrió un error al crear el excel intermedio: {str(e)}')
        return None
# Función para filtrar cambios del archivo Excel
def filtros_excel(file_path):

    try:
        # Cargar el archivo Excel copiado
        workbook = load_workbook(filename=file_path)

        # Hoja activa
        sheet = workbook.active

        # Rango de celdas de la hoja
        max_row = sheet.max_row
        max_column = sheet.max_column

        # Listas para almacenar las filas con errores
        filas_sin_elementos = set()
        filas_con_errores = set()
        elementos_no_encontrados = set()
        interno_error = set()

        # 1. Detectar filas con 4 o más celdas vacías en columnas específicas
        for row in range(1, max_row + 1):
            empty_cell_count = 0

            for column in ['A', 'B', 'C', 'D', 'L', 'M', 'N', 'O']:
                cell = sheet[f"{column}{row}"]
                
                if cell.value is None:
                    empty_cell_count += 1

                if empty_cell_count >= 4:
                    filas_sin_elementos.add(row)
                    break  # Salir del bucle si hay 4 o más celdas vacías

        ## 2. Eliminar -W en columnas D y O (si no coincide, a errores)
        for row in range(1, sheet.max_row + 1):

            # Procesar la columna D
            cell_value_D = sheet.cell(row=row, column=4).value

            if cell_value_D:  # Verificar si hay un valor en la columna D
                values_D = cell_value_D.split()  # Dividir los valores de la celda en la columna D

                if values_D:
                    first_value_D = None  # Variable temporal para almacenar el primer valor

                    for value in values_D:
                        if not value.startswith("-W"):
                            first_value_D = value
                            break

                    if first_value_D is not None:
                        sheet.cell(row=row, column=4).value = first_value_D  # Actualizar la celda en la columna D solo si es necesario
                    else:
                        elementos_no_encontrados.add(row)

            # Procesar la columna O
            cell_value_O = sheet.cell(row=row, column=15).value

            if cell_value_O:  # Verificar si hay un valor en la columna O
                values_O = cell_value_O.split() # Dividir los valores de la celda en la columna O

                if values_O:
                    first_value_O = None  # Variable temporal para almacenar el primer valor

                    for value in values_O:
                        if not value.startswith("-W"):
                            first_value_O = value
                            break

                    if first_value_O is not None:
                        sheet.cell(row=row, column=15).value = first_value_O  # Actualizar la celda en la columna O solo si es necesario
                    else:
                        elementos_no_encontrados.add(row)

        
        ## 3. Eliminar duplicados en las columnas A, B, C, D y L, M, N, O
        for column_letter in ['A', 'B', 'C', 'D', 'L', 'M', 'N', 'O']:

            for row in range(1, sheet.max_row + 1):
                value_column = sheet.cell(row=row, column=column_index_from_string(column_letter)).value

                if value_column:
                    values = value_column.split()
                    unique_values = []

                    for value in values:
                        if value not in unique_values:
                            unique_values.append(value)

                    sheet.cell(row=row, column=column_index_from_string(column_letter)).value = ' '.join(unique_values)


        ## 4. Comprobar contenido de las celdas de las columnas B, C, D, M, N, O con la columna H (si no, a errores)
        for row in range(1, sheet.max_row + 1):
            value_column_H = sheet.cell(row=row, column=8).value
            if value_column_H and "/" in value_column_H:
                split_values_H = value_column_H.split("/")

                # Procesar la columna B
                value_column_B = sheet.cell(row=row, column=2).value
                if value_column_B:
                    updated_B = False
                    for subvalue in value_column_B.split():
                        if subvalue in split_values_H[0]:
                            sheet.cell(row=row, column=2).value = subvalue
                            updated_B = True
                            break
                    if not updated_B:
                        filas_con_errores.add(row)

                # Procesar la columna C
                value_column_C = sheet.cell(row=row, column=3).value
                if value_column_C:
                    updated_C = False
                    for subvalue in value_column_C.split():
                        if subvalue in split_values_H[0]:
                            sheet.cell(row=row, column=3).value = subvalue
                            updated_C = True
                            break
                    if not updated_C:
                        filas_con_errores.add(row)

                # Procesar la columna D
                value_column_D = sheet.cell(row=row, column=4).value
                if value_column_D:
                    updated_D = False
                    for subvalue in value_column_D.split():
                        if subvalue in split_values_H[0]:
                            sheet.cell(row=row, column=4).value = subvalue
                            updated_D = True
                            break
                    if not updated_D:
                        filas_con_errores.add(row)

                # Procesar la columna M
                value_column_M = sheet.cell(row=row, column=13).value
                if value_column_M:
                    updated_M = False
                    for subvalue in value_column_M.split():
                        if subvalue in split_values_H[1]:
                            sheet.cell(row=row, column=13).value = subvalue
                            updated_M = True
                            break
                    if not updated_M:
                        filas_con_errores.add(row)

                # Procesar la columna N
                value_column_N = sheet.cell(row=row, column=14).value
                if value_column_N:
                    updated_N = False
                    for subvalue in value_column_N.split():
                        if subvalue in split_values_H[1]:
                            sheet.cell(row=row, column=14).value = subvalue
                            updated_N = True
                            break
                    if not updated_N:
                        filas_con_errores.add(row)

                # Procesar la columna O
                value_column_O = sheet.cell(row=row, column=15).value
                if value_column_O:
                    updated_O = False
                    for subvalue in value_column_O.split():
                        if subvalue in split_values_H[1]:
                            sheet.cell(row=row, column=15).value = subvalue
                            updated_O = True
                            break
                    if not updated_O:
                        filas_con_errores.add(row)

            else:
                filas_con_errores.add(row)

        ## 5. Comprobar si las columnas B y M tienen el mismo valor "CC, CP u OP"
        for row in range(1, max_row + 1):

            # Obtener los valores de las columnas B y M
            valor_columna_B = sheet.cell(row=row, column=2).value
            valor_columna_M = sheet.cell(row=row, column=13).value
            
            # Verificar si ambos valores no están vacíos y cumplen con los requisitos
            if valor_columna_B is not None and valor_columna_M is not None:

                # Eliminar espacios en blanco alrededor de los valores y convertir a mayúsculas para comparar
                valor_columna_B = valor_columna_B.strip().upper()
                valor_columna_M = valor_columna_M.strip().upper()
                
                # Verificar si los valores comienzan con "CC", "CP" u "OP" y son iguales
                if valor_columna_B.startswith(("++CC", "++CP", "++OP")) and valor_columna_M.startswith(("++CC", "++CP", "++OP")) and valor_columna_B == valor_columna_M:

                    # Añadir la fila al conjunto de filas con errores
                    interno_error.add(row)

        ## 6. Agregar saltos de línea en la columna H
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=8)  # Columna H
            value = cell.value

            if value and isinstance(value, str):
                # Reemplazar "/ " por "/\n"
                new_value = value.replace("/", "/\n")

                # Asignar el nuevo valor a la celda
                cell.value = new_value

        #  (corregir)
        #  Eliminar el "=" de las celdas en columnas A, H, K, L
        columns_to_process = [1, 8, 11, 12]   

        for row in range(1, sheet.max_row + 1):
            
            for col in columns_to_process:
                cell_value = sheet.cell(row=row, column=col).value

                if isinstance(cell_value, str) and cell_value.startswith('='):
                    sheet.cell(row=row, column=col).value = cell_value.lstrip('=')

        # Crear hoja llamada 'errores' si no existe
        if 'errores' not in workbook.sheetnames:
            error_sheet = workbook.create_sheet('errores')

            # Agregar las cabeceras de las columnas
            error_sheet['A1'] = 'Nombre de la Manguera:'
            error_sheet['B1'] = 'Motivo del error:'
            error_sheet['C1'] = 'Página dónde se encuentra:'

        else:
            error_sheet = workbook['errores']

        # Mover las filas con errores a la hoja 'errores' (solo nombre Manguera)
        error_row_idx = error_sheet.max_row + 1  # Empezar después de la última fila en la hoja 'errores'

        for row_index in filas_sin_elementos.union(filas_con_errores, elementos_no_encontrados, interno_error):

            # Obtener el valor de la columna 'Manguera' de la fila con error
            manguera_value = sheet.cell(row=row_index, column=8).value

            # Determinar el tipo de error y asignar el mensaje correspondiente
            if row_index in filas_sin_elementos:
                motivo_error = "Manguera sin Origenes ni Destinos conectados"

            elif row_index in filas_con_errores:
                motivo_error = '''El Origen y/o Destino no se 
encuentra en el nombre
 de la Manguera'''
    
            elif row_index in elementos_no_encontrados:
                motivo_error = "Elementos no encontrados en el nombre de la Manguera"

            elif row_index in interno_error:
                value_column_B = sheet.cell(row=row_index, column=2).value
                motivo_error = f"Manguera interna {value_column_B}"

            else:
                motivo_error = "Error desconocido"

            # Copiar el valor de la columna 'Manguera' a la hoja 'errores'
            error_sheet.cell(row=error_row_idx, column=1, value=manguera_value)

            # Agregar la página donde se encuentra el error
            page_value = sheet.cell(row=row_index, column=11).value  # Obtener el valor de la columna K (columna 11)
            error_sheet.cell(row=error_row_idx, column=2, value=page_value)

            # Agregar el motivo del error en la columna 'Motivo del error' de la hoja 'errores'
            error_sheet.cell(row=error_row_idx, column=3, value=motivo_error)

            error_row_idx += 1

        # Eliminar las filas que cumplen las condiciones en la hoja original
        for row_index in sorted(filas_sin_elementos.union(filas_con_errores, interno_error), reverse=True):
            sheet.delete_rows(row_index)

        print(f'Los filtros han sido guardados con éxito en {file_path}, puede continuar para generar los Diagramas Unifilares.')
        print() 

        # Guardar los cambios en el archivo Excel copiado
        workbook.save(filename=file_path)

    except Exception as e:
        print(f'Ocurrió un error al filtrar los datos del Excel: {str(e)}')

def filtros_uniones(file_path):
    try:
        # Cargar el archivo Excel copiado
        workbook = load_workbook(filename=file_path)

        # Obtener la hoja activa y las hojas 'uniones'
        sheet = workbook.active
        
        if 'uniones' not in workbook.sheetnames:
            union_sheet = workbook.create_sheet('uniones')
        else:
            union_sheet = workbook['uniones']

        # Inicializar el índice de la fila para la hoja 'uniones'
        union_row_idx = 1

        # Iterar por las filas de la hoja activa
        for row in range(1, sheet.max_row + 1):

            # Unir los valores de las columnas A, B, C, D, F, G en una sola cadena
            value_A = sheet.cell(row=row, column=1).value or ""
            value_B = sheet.cell(row=row, column=2).value or ""
            value_C = sheet.cell(row=row, column=3).value or ""

            # Tener en cuenta la parte izquierda del punto en D
            value_D = sheet.cell(row=row, column=4).value or ""
            if '.' in value_D:
                value_D = value_D.split('.')[0]

            value_F = sheet.cell(row=row, column=6).value or ""
            value_G = sheet.cell(row=row, column=7).value or ""

            # Unir los valores de las columnas H, I, y J en una sola cadena
            value_H = sheet.cell(row=row, column=8).value or ""
            value_I = sheet.cell(row=row, column=9).value or ""
            value_J = sheet.cell(row=row, column=10).value or ""

            # Unir los valores de las columnas L, M, N, O, Q, R en una sola cadena
            value_L = sheet.cell(row=row, column=12).value or ""
            value_M = sheet.cell(row=row, column=13).value or ""
            value_N = sheet.cell(row=row, column=14).value or ""

            # Tener en cuenta la parte izquierda del punto en O:
            value_O = sheet.cell(row=row, column=15).value or ""
            if '.' in value_O:
                value_O = value_O.split('.')[0]

            value_Q = sheet.cell(row=row, column=17).value or ""
            value_R = sheet.cell(row=row, column=18).value or ""

            # Escribir los valores en la hoja 'uniones' en las columnas A, B y C respectivamente
            union_sheet.cell(row=union_row_idx, column=1).value = f"{value_A}{value_B} \n {value_C}{value_D} \n {value_F} \n {value_G}"
            union_sheet.cell(row=union_row_idx, column=2).value = f"{value_H} \n {value_I} \n {value_J}"
            union_sheet.cell(row=union_row_idx, column=3).value = f"{value_L}{value_M} \n {value_N}{value_O} \n {value_Q} \n {value_R}"

            union_row_idx += 1

        ## Union de las filas ##
        data = [] # Leer todas las filas en la hoja 'uniones'

        for row in union_sheet.iter_rows(values_only=True):
            data.append(list(row))

        # Inicializar una nueva lista para almacenar las filas unidas
        merged_data = []

        # Realizar las uniones según las reglas
        i = 0
        while i < len(data): # Mientras haya fila de datos

            current_row = data[i] # Fila actual con indice i
            j = i + 1 # Fila siguiente, variable j

            while j < len(data):
                next_row = data[j] # Selecciona siguiente fila para comparar a la actual

                if current_row[-1] == next_row[0]:
                    current_row.extend(next_row[1:])
                    data.pop(j)

                elif current_row[-1] == next_row[-1]:

                    # Concatenar las filas de izquierda a derecha
                    merged_row = current_row + next_row[1:]  # Se omite el primer elemento de next_row
                    data[i] = merged_row  # Reemplazar current_row con la fila concatenada
                    data.pop(j)

                else: 
                    j += 1

            merged_data.append(current_row)
            i += 1

        # Ordenar por longitud de filas descendente
        merged_data.sort(key=lambda x: len(x), reverse=True)

        # Limpiar cualquier dato existente en la hoja 'uniones'
        union_sheet.delete_rows(1, union_sheet.max_row)

        # Escribir las filas unidas ordenadas en la hoja 'uniones'
        for row in merged_data:
            union_sheet.append(row)

        # Guardar los cambios en el archivo Excel copiado
        workbook.save(filename=file_path)

    except Exception as e:
        print(f'Ocurrió un error al crear las uniones del Excel: {str(e)}')


def filtros_mangueras_individuales(file_path):
    try:

        # Cargar el archivo Excel
        workbook = load_workbook(filename=file_path)

        # Obtener la hoja 'uniones' y crear la hoja 'hoja_mangueras_individuales' si no existe
        union_sheet = workbook['uniones']

        if 'mangueras_individuales' not in workbook.sheetnames:
            mangueras_sheet = workbook.create_sheet('mangueras_individuales')

        else:
            mangueras_sheet = workbook['mangueras_individuales']

        
        mangueras_row_idx = 1 # Contador para el índice de la fila en 'hoja_mangueras_individuales'
        row_idx = 1 # Iterar por las filas de la hoja 'uniones' y seleccionar aquellas con máximo 3 columnas

        while row_idx <= union_sheet.max_row:
            row = union_sheet[row_idx]

            # Contar el número de celdas con valor en la fila
            count_cells = sum(1 for cell in row if cell.value is not None)

            if count_cells == 3:

                # Copiar la fila a la hoja 'hoja_mangueras_individuales'
                for cell in row:
                    mangueras_sheet.cell(row=mangueras_row_idx, column=cell.column, value=cell.value)
                mangueras_row_idx += 1

                # Eliminar la fila de 'uniones'
                union_sheet.delete_rows(row_idx, 1)
                continue

            row_idx += 1

        # Guardar los cambios en el archivo Excel
        workbook.save(filename=file_path)

    except Exception as e:
        print(f'Ocurrió un error al procesar el archivo Excel: {str(e)}')


# Función para leer los datos de la hoja 'errores'
def leer_datos_errores():

    global new_file_path  # Asegurar que estamos utilizando la variable global

    try:
        # Verificar que new_file_path no sea None
        if not new_file_path:
            messagebox.showinfo("Información", "No se ha seleccionado ningún archivo para procesar.")
            return []

        # Cargar el archivo Excel
        workbook = load_workbook(filename=new_file_path)
        if 'errores' not in workbook.sheetnames:
            messagebox.showinfo("Información", "La hoja 'errores' no existe en el archivo.")
            return []

        sheet = workbook['errores']

        # Obtener los datos de la hoja 'errores'
        datos_errores = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            datos_errores.append(row)

        return datos_errores

    except Exception as e:
        print(f"Ocurrió un error al leer los datos de la hoja 'errores': {str(e)}")
        return []

# Función para procesar el archivo seleccionado
def process_file():

    global new_file_path  # Asegurar que estamos utilizando la variable global

    if not new_file_path:
        print("No se seleccionó ningún archivo. El programa se cerrará.")
        return

    tamaño_hoja = size_var.get()
    figsize, x_position_max = get_figsize_and_max_pos(tamaño_hoja)

    # Crear un objeto para combinar PDFs
    pdf_merger = PdfMerger()

    # Abrir el archivo de Excel con xlwings
    app = xw.App(visible=False)
    app.display_alerts = False
    app.screen_updating = False

    try:

        wb = app.books.open(new_file_path)  # Utilizar new_file_path

        ### Procesar la hoja 'uniones'
        sheet_uniones = wb.sheets['uniones']
        df_uniones = sheet_uniones.used_range.value

        # Iterar sobre las filas del DataFrame
        for idx, row in enumerate(df_uniones, start=1):

            # Verificar si la fila tiene suficientes valores para dibujar el gráfico
            if len(row) > 1:

                # Crear un nuevo grafo para cada fila
                G = create_graph_from_row(row)

                # Dibujar el gráfico con el tamaño adecuado
                fig, ax = plt.subplots(figsize=figsize)

                # Calcular posiciones de los nodos
                pos = generate_positions(G, x_position_max, tamaño_hoja)
                draw_graph(G, pos, ax)

                # Guardar la figura en un objeto BytesIO
                buf = BytesIO()
                plt.savefig(buf, format='pdf')
                buf.seek(0)
                pdf_merger.append(buf)
                buf.close()  # Cerrar BytesIO para liberar memoria
                plt.close(fig)  # Cerrar la figura para liberar memoria

                print(f"El diagrama de unifilares para la fila {idx} se ha generado con éxito\n")
            else:
                print(f"No ha sido posible hacer un diagrama para la fila {idx} porque no hay suficientes datos.\n")

         ### Leer datos de la hoja 'mangueras_individuales'
        sheet_mangueras = wb.sheets['mangueras_individuales']
        df_mangueras = sheet_mangueras.used_range.value

        # Preparar los datos para la tabla
        data = [['Elemento Origen', 'Manguera', 'Elemento Destino']]
        for idx, row in enumerate(df_mangueras, start=1):
            if len(row) == 3:
                data.append([row[0], row[1], row[2]])
            else:
                print(f"Advertencia: La fila {idx} no tiene tres conexiones. Se omitirá.")
        
        # Configurar el estilo de tabla
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.skyblue),  # Color de fondo para encabezado
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alinear todo al centro
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Color de texto para encabezado
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Borde interior de celdas
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Borde exterior de tabla
        ])

        # Crear la tabla con los datos y estilo configurados
        table = Table(data)
        table.setStyle(style)

        # Crear un objeto BytesIO para guardar el PDF de la tabla
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        elements = []
        elements.append(table)
        doc.build(elements)

        # Añadir el PDF de la tabla al pdf_merger
        buf.seek(0)
        pdf_merger.append(buf)
        buf.close()

        ### Configurar la hoja 'errores' para PDF
        sheet_errores = wb.sheets['errores']

        sheet_errores.page_setup.paperSize = 9  # Código numérico para tamaño A4
        sheet_errores.page_setup.fitToWidth = 1
        sheet_errores.page_setup.fitToHeight = False

        sheet_errores.left_margin = 0.5
        sheet_errores.right_margin = 0.5
        sheet_errores.top_margin = 0.5
        sheet_errores.bottom_margin = 0.5
        sheet_errores.header_margin = 0.5
        sheet_errores.footer_margin = 0.5

        # Crear un objeto BytesIO para guardar el PDF de la tabla de errores
        buf_errores = BytesIO()
        
        # Leer datos de la hoja 'errores'
        datos_errores = sheet_errores.used_range.value[1:]  # Excluir el encabezado
        
        # Configurar estilo de la tabla
        style_errores = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.skyblue),  # Color de fondo para encabezado
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alinear todo al centro
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Color de texto para encabezado
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Borde interior de celdas
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Borde exterior de tabla
        ])

        # Añadir títulos a los datos de errores
        datos_errores.insert(0, ['Nombre de la Manguera', 'Página donde se encuentra', 'Motivo del error'])

        # Crear la tabla de errores con los datos y estilo configurados
        table_errores = Table(datos_errores)
        table_errores.setStyle(style_errores)

        # Crear un objeto BytesIO para guardar el PDF de la tabla de errores
        buf_errores = BytesIO()
        doc_errores = SimpleDocTemplate(buf_errores, pagesize=A4)
        
        elements_errores = []
        elements_errores.append(table_errores)
        doc_errores.build(elements_errores)

        # Añadir el PDF de la tabla de errores al pdf_merger
        buf_errores.seek(0)
        pdf_merger.append(buf_errores)
        buf_errores.close()

        # Guardar el archivo PDF combinado
        save_combined_pdf(pdf_merger)

    except Exception as e:
        print(f"Ocurrió un error al procesar el archivo: {str(e)}")

    finally:

        # Cerrar el libro de Excel y la aplicación
        wb.close()
        app.quit()


# Función para obtener el tamaño de la figura y la posición máxima en X
def get_figsize_and_max_pos(tamaño_hoja):
    
    if tamaño_hoja == 'A4': 
        return (210 / 25.4, 297 / 25.4), 4 # Tamaño A4
    else:                   
        return (420 / 25.4, 297 / 25.4), 8 # Tamaño A3


# Función para crear un grafo a partir de una fila
def create_graph_from_row(row):

    # Crear un nuevo grafo G
    G = nx.Graph() 
    
    # Agregar nodos al grafo G
    for cell in row:
        if cell is not None:  # Verificar si el valor de la celda no es None
            G.add_node(cell)

    # Agregar conexiones entre nodos en la fila actual
    for i in range(len(row) - 1):
        if row[i] is not None and row[i + 1] is not None:  # Verificar que los valores no son None
            G.add_edge(row[i], row[i + 1])
    return G


# Función para generar posiciones de los nodos
def generate_positions(G, x_position_max, tamaño_hoja):
    pos = {}
    x_position, y_position = 0, 0
    max_nodes_per_row = 3 if tamaño_hoja == 'A4' else 5

    for node in G.nodes():
        pos[node] = (x_position, y_position)

        if (x_position + 1) % max_nodes_per_row == 0:
            x_position = 1  # Empezar desde 1 en la siguiente fila
            y_position -= 1
        else:
            x_position += 1

    return pos

# Función para dibujar los nodos del grafo
def draw_graph(G, pos, ax):

    node_colors = ['skyblue' if i % 2 == 0 else 'lightgreen' for i in range(len(G.nodes))]
    node_shapes = ['s' if i % 2 == 0 else 'd' for i in range(len(G.nodes))]

    for i, (node, (x, y)) in enumerate(pos.items()):

        # Si el nodo es un cuadrado azul y no tiene un '=', se añade
        if node_shapes[i] == 's' and node_colors[i] == 'skyblue' and not str(node).startswith('='):
            labeled_node = '=' + str(node)

        else:
            labeled_node = str(node)

        # Dibujar el nodo con sus atributos
        nx.draw_networkx_nodes(G, pos, nodelist=[node], node_size=3000, node_shape=node_shapes[i], node_color=node_colors[i])
        
        # Añadir el texto del nodo, usando labeled_node con tamaño de fuente ajustado
        nx.draw_networkx_labels(G, pos, labels={node: labeled_node}, font_size=6, ax=ax)

    nx.draw_networkx_edges(G, pos)
    ax.axis('off')


# Función para guardar el archivo PDF combinado
def save_combined_pdf(pdf_merger):
    pdf_combined_file = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
    
    if not pdf_combined_file:
        messagebox.show_warning("No se seleccionó ninguna ruta para guardar el archivo. El programa se cerrará.") # El programa se cierra si no se selecciona una ruta
        root.destroy()
    else:
        with open(pdf_combined_file, 'wb') as output_pdf:
            pdf_merger.write(output_pdf)
        pdf_merger.close()  # Cerrar el PdfMerger para liberar recursos

        # Mensaje de Información al finalizar el proceso
        messagebox.showinfo("Información", f"El archivo PDF se ha generado correctamente en: {pdf_combined_file}\n")


###### Configuración de la apariencia de la ventana principal con customtkinter #####

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

root = ctk.CTk() # Crear la ventana principal
root.title("Generador de Diagramas Unifilares")
root.resizable(False, False)  # Desactivar redimensionamiento

# Botón 'Examinar'
ctk.CTkLabel(root, text="Seleccionar archivo del Excel:").grid(row=0, column=0, padx=10, pady=10)
file_entry = ctk.CTkEntry(root, width=300)
file_entry.grid(row=0, column=1, padx=10, pady=10)
ctk.CTkButton(root, text="Examinar", command=select_file).grid(row=0, column=2, padx=10, pady=10)

# Botón 'Generar PDF'
ctk.CTkButton(root, text="Generar PDF", command=process_file).grid(row=3, column=0, columnspan=3, padx=10, pady=10) 

# Botón 'Tamaño hoja'
ctk.CTkLabel(root, text="Seleccionar tamaño de hoja:").grid(row=1, column=0, padx=10, pady=10)
size_var = ctk.StringVar(value="A4")
size_combobox = ctk.CTkComboBox(root, variable=size_var, values=["A4", "A3"])
size_combobox.grid(row=1, column=1, padx=10, pady=10)
 
# Botón 'Salir'
ctk.CTkButton(root, text="Salir", command=root.quit).grid(row=4, column=0, columnspan=3, pady=10)

# Cuadro de texto
text_widget = ctk.CTkTextbox(root, wrap='word', height=200, width=600)
text_widget.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

sys.stdout = RedirectStdout(text_widget)

# Mantener la ventana abierta y a la espera de eventos (clics de ratón, pulsaciones de teclas, etc.)
root.mainloop()
